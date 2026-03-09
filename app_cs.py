#!/usr/bin/env python3
"""
Script ContentSquare - Conversions E-commerce (VERSION DOCUMENTÉE)
Utilise les endpoints officiels de la Metrics API ContentSquare
"""

import os
import json
import requests
from dotenv import load_dotenv
from datetime import datetime, timezone
from pathlib import Path
from requests.adapters import HTTPAdapter
from requests.exceptions import RequestException, SSLError, Timeout
from urllib3.util.retry import Retry
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python3 -m pip install openpyxl"
    ) from exc

try:
    from contentsquare_config import PAGE_GROUP_MAPPING_ID as CONFIG_PAGE_GROUP_MAPPING_ID
except Exception:
    CONFIG_PAGE_GROUP_MAPPING_ID = 2066672
try:
    from contentsquare_config import PAGE_GROUP_ID as CONFIG_PAGE_GROUP_ID
except Exception:
    CONFIG_PAGE_GROUP_ID = None
try:
    from contentsquare_config import GOAL_IDS as CONFIG_GOAL_IDS
except Exception:
    CONFIG_GOAL_IDS = []
try:
    from contentsquare_config import SEGMENT_IDS_TO_ANALYZE as CONFIG_SEGMENT_IDS_TO_ANALYZE
except Exception:
    CONFIG_SEGMENT_IDS_TO_ANALYZE = []
try:
    from contentsquare_config import ANALYZE_BY_DEVICE as CONFIG_ANALYZE_BY_DEVICE
except Exception:
    CONFIG_ANALYZE_BY_DEVICE = True
try:
    from contentsquare_config import START_DATE as CONFIG_START_DATE
except Exception:
    CONFIG_START_DATE = None
try:
    from contentsquare_config import END_DATE as CONFIG_END_DATE
except Exception:
    CONFIG_END_DATE = None

# Charger les identifiants
load_dotenv()
CLIENT_ID = os.getenv("CS_CLIENT_ID")
CLIENT_SECRET = os.getenv("CS_CLIENT_SECRET")
PROJECT_ID = os.getenv("CS_PROJECT_ID")

SEGMENT_IDS_TO_ANALYZE = []
ANALYZE_BY_DEVICE = True
START_DATE = CONFIG_START_DATE
END_DATE = CONFIG_END_DATE
GOAL_ID = None
EXPORT_DIR = "exports"
KPI_EXCEL_FILENAME = "contentsquare_kpis.xlsx"
REQUEST_TIMEOUT_SECONDS = float(os.getenv("CS_REQUEST_TIMEOUT_SECONDS", "30"))
RETRY_TOTAL = int(os.getenv("CS_RETRY_TOTAL", "3"))


def resolve_mapping_id(value, default=2066672):
    if value is None:
        return default
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def resolve_optional_int(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def resolve_goal_ids(value):
    if value is None:
        return []

    raw_ids = value if isinstance(value, (list, tuple, set)) else [value]
    result = []
    for item in raw_ids:
        resolved = resolve_optional_int(item)
        if resolved is not None:
            result.append(resolved)
    return result


def resolve_segment_ids(value):
    return resolve_goal_ids(value)


def resolve_bool(value, default=True):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "1", "yes", "y", "on"}:
            return True
        if normalized in {"false", "0", "no", "n", "off"}:
            return False
    return default


def parse_config_date(value, field_name):
    if value is None:
        raise ValueError(f"{field_name} est manquant. Définis-le dans contentsquare_config.py.")
    value_str = str(value).strip()
    try:
        return datetime.strptime(value_str, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(
            f"{field_name} invalide ({value_str}). Format attendu: YYYY-MM-DD."
        ) from exc


PAGE_GROUP_MAPPING_ID = resolve_mapping_id(CONFIG_PAGE_GROUP_MAPPING_ID)
PAGE_GROUP_ID = resolve_optional_int(CONFIG_PAGE_GROUP_ID)
GOAL_IDS = resolve_goal_ids(CONFIG_GOAL_IDS)
SEGMENT_IDS_TO_ANALYZE = resolve_segment_ids(CONFIG_SEGMENT_IDS_TO_ANALYZE)
ANALYZE_BY_DEVICE = resolve_bool(CONFIG_ANALYZE_BY_DEVICE, default=True)


def get_analysis_devices(analyze_by_device):
    if analyze_by_device:
        return ["all", "desktop", "tablet", "mobile"]
    return ["all"]


def build_http_session():
    retry = Retry(
        total=RETRY_TOTAL,
        connect=RETRY_TOTAL,
        read=RETRY_TOTAL,
        status=RETRY_TOTAL,
        backoff_factor=0.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset({"GET", "POST"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


HTTP_SESSION = build_http_session()


def request_json(method, url, headers=None, params=None, json=None):
    try:
        response = HTTP_SESSION.request(
            method=method,
            url=url,
            headers=headers,
            params=params,
            json=json,
            timeout=REQUEST_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        return response.json()
    except Timeout as exc:
        raise RuntimeError(
            f"Timeout après {REQUEST_TIMEOUT_SECONDS}s vers {url}. "
            "Vérifie la connectivité réseau et la valeur du endpoint."
        ) from exc
    except SSLError as exc:
        raise RuntimeError(
            f"Erreur SSL vers {url}. Vérifie les certificats locaux, proxy/VPN ou endpoint."
        ) from exc
    except RequestException as exc:
        status_code = getattr(exc.response, "status_code", None)
        if status_code:
            raise RuntimeError(f"HTTP {status_code} sur {url}: {exc}") from exc
        raise RuntimeError(f"Erreur réseau vers {url}: {exc}") from exc


def get_token(client_id, client_secret, project_id):
    auth_url = "https://api.eu-west-1.production.contentsquare.com/v1/oauth/token"
    payload = {
        "client_id": client_id,
        "client_secret": client_secret,
        "grant_type": "client_credentials",
        "scope": "metrics",
        "projectId": project_id
    }
    data = request_json("POST", auth_url, json=payload)
    token = data.get("access_token")
    endpoint = data.get("endpoint")
    if not token or not endpoint:
        raise RuntimeError("Réponse OAuth invalide: access_token ou endpoint manquant.")
    return token, endpoint


def get_segments(endpoint, token, project_id):
    url = f"{endpoint}/v1/segments"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"projectId": project_id}
    return request_json("GET", url, headers=headers, params=params)


def get_goals(endpoint, token, project_id):
    url = f"{endpoint}/v1/goals"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"projectId": project_id}
    return request_json("GET", url, headers=headers, params=params)


def get_mappings(endpoint, token, project_id):
    url = f"{endpoint}/v1/mappings"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"projectId": project_id}
    return request_json("GET", url, headers=headers, params=params)


def get_mapping_page_groups(endpoint, token, project_id, mapping_id):
    url = f"{endpoint}/v1/mappings/{mapping_id}/page-groups"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"projectId": project_id}
    return request_json("GET", url, headers=headers, params=params)


def get_page_groups_for_mapping(endpoint, token, project_id, mapping_id):
    mappings_data = get_mappings(endpoint, token, project_id)
    mappings = mappings_data.get("payload", [])
    mapping_name = ""
    for mapping in mappings:
        if str(mapping.get("id")) == str(mapping_id):
            mapping_name = mapping.get("name", "")
            break

    page_groups_data = get_mapping_page_groups(endpoint, token, project_id, mapping_id)
    page_groups = page_groups_data.get("payload", [])

    result = []
    for page_group in page_groups:
        page_group_id = page_group.get("id")
        if page_group_id is None:
            continue
        result.append(
            {
                "id": page_group_id,
                "name": page_group.get("name"),
                "category": page_group.get("category"),
                "mapping_id": mapping_id,
                "mapping_name": mapping_name,
            }
        )

    return sorted(result, key=lambda item: item["id"])


def get_all_page_groups(endpoint, token, project_id):
    mappings_data = get_mappings(endpoint, token, project_id)
    mappings = mappings_data.get("payload", [])
    all_groups = []

    for mapping in mappings:
        mapping_id = mapping.get("id")
        if mapping_id is None:
            continue

        page_groups_data = get_mapping_page_groups(endpoint, token, project_id, mapping_id)
        for page_group in page_groups_data.get("payload", []):
            page_group_id = page_group.get("id")
            if page_group_id is None:
                continue
            all_groups.append(
                {
                    "id": page_group_id,
                    "name": page_group.get("name"),
                    "category": page_group.get("category"),
                    "mapping_id": mapping_id,
                    "mapping_name": mapping.get("name", ""),
                }
            )

    return sorted(all_groups, key=lambda item: (item["mapping_id"], item["id"]))


def find_page_group_by_id(endpoint, token, project_id, page_group_id):
    target_id = str(page_group_id)
    mappings_data = get_mappings(endpoint, token, project_id)
    mappings = mappings_data.get("payload", [])

    for mapping in mappings:
        mapping_id = mapping.get("id")
        if mapping_id is None:
            continue
        page_groups_data = get_mapping_page_groups(endpoint, token, project_id, mapping_id)
        for page_group in page_groups_data.get("payload", []):
            if str(page_group.get("id")) == target_id:
                return {
                    "id": page_group.get("id"),
                    "name": page_group.get("name"),
                    "category": page_group.get("category"),
                    "mapping_id": mapping_id,
                    "mapping_name": mapping.get("name", ""),
                }
    return None


def get_page_group_metrics(endpoint, token, project_id, page_group_id, start_date, end_date, device="all", segment_ids=None):
    url = f"{endpoint}/v1/metrics/page-group/{page_group_id}"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "projectId": project_id,
        "startDate": start_date,
        "endDate": end_date,
        "device": device,
    }
    if segment_ids:
        params["segmentIds"] = ",".join(map(str, segment_ids))
    return request_json("GET", url, headers=headers, params=params)


def get_page_group_web_vitals(endpoint, token, project_id, page_group_id, start_date, end_date, device="all", segment_ids=None):
    url = f"{endpoint}/v1/metrics/page-group/{page_group_id}/web-vitals"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "projectId": project_id,
        "startDate": start_date,
        "endDate": end_date,
        "device": device,
    }
    if segment_ids:
        params["segmentIds"] = ",".join(map(str, segment_ids))
    return request_json("GET", url, headers=headers, params=params)


def get_page_group_conversion_rate(endpoint, token, project_id, page_group_id, start_date, end_date, goal_id, device="all", segment_ids=None):
    url = f"{endpoint}/v1/metrics/page-group/{page_group_id}/conversion-rate"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "projectId": project_id,
        "startDate": start_date,
        "endDate": end_date,
        "device": device,
        "goalId": goal_id,
    }
    if segment_ids:
        params["segmentIds"] = ",".join(map(str, segment_ids))
    return request_json("GET", url, headers=headers, params=params)


def get_site_metrics(endpoint, token, project_id, start_date, end_date, device="all", segment_ids=None):
    url = f"{endpoint}/v1/metrics/site"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "projectId": project_id,
        "startDate": start_date,
        "endDate": end_date,
        "device": device
    }
    if segment_ids:
        params["segmentIds"] = ",".join(map(str, segment_ids))
    return request_json("GET", url, headers=headers, params=params)


def get_ecommerce_conversions(endpoint, token, project_id, start_date, end_date, goal_id=None, device="all", segment_ids=None):
    url = f"{endpoint}/v1/metrics/site/conversions"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "projectId": project_id,
        "startDate": start_date,
        "endDate": end_date,
        "device": device
    }
    if goal_id:
        params["goalId"] = goal_id
    if segment_ids:
        params["segmentIds"] = ",".join(map(str, segment_ids))
    return request_json("GET", url, headers=headers, params=params)


def get_ecommerce_conversion_rate(endpoint, token, project_id, start_date, end_date, goal_id=None, device="all", segment_ids=None):
    url = f"{endpoint}/v1/metrics/site/conversion-rate"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "projectId": project_id,
        "startDate": start_date,
        "endDate": end_date,
        "device": device
    }
    if goal_id:
        params["goalId"] = goal_id
    if segment_ids:
        params["segmentIds"] = ",".join(map(str, segment_ids))
    return request_json("GET", url, headers=headers, params=params)


def extract_metric_value(metrics_data, metric_name, fallback_key=None):
    payload = metrics_data.get("payload", {}) if isinstance(metrics_data, dict) else {}

    # Nouveau format API: payload.values = [{name: "...", value: ...}, ...]
    values = payload.get("values")
    if isinstance(values, list):
        for item in values:
            if isinstance(item, dict) and item.get("name") == metric_name:
                return item.get("value")

    # Ancien format API: payload.<key>
    if fallback_key:
        return payload.get(fallback_key)
    return payload.get(metric_name)


def extract_single_value(metrics_data, preferred_name=None):
    payload = metrics_data.get("payload", {}) if isinstance(metrics_data, dict) else {}

    value = payload.get("value")
    if isinstance(value, (int, float)):
        return value

    values = payload.get("values")
    if isinstance(values, list) and values:
        if preferred_name:
            for item in values:
                if isinstance(item, dict) and item.get("name") == preferred_name:
                    return item.get("value", 0)
        if isinstance(values[0], dict):
            return values[0].get("value", 0)
    return 0


def format_count(value):
    if not isinstance(value, (int, float)):
        return "N/A"
    return f"{int(round(value)):,}"


def format_percentage(value):
    if not isinstance(value, (int, float)):
        return "N/A"
    return f"{value:.2f}%"


def export_ids_file(export_dir, filename_prefix, label, rows):
    export_path = Path(export_dir)
    export_path.mkdir(parents=True, exist_ok=True)

    file_path = export_path / f"{filename_prefix}.txt"
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    with file_path.open("w", encoding="utf-8") as f:
        f.write(f"# {label}\n")
        f.write(f"# generated_at_utc: {generated_at}\n")
        f.write(f"# total: {len(rows)}\n")
        f.write("# format: id\\tname\\textra\n")
        for row in rows:
            f.write(f"{row.get('id', '')}\t{row.get('name', '')}\t{row.get('extra', '')}\n")

    return file_path


def normalize_excel_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def metrics_response_to_rows(metrics_data):
    payload = metrics_data.get("payload", {}) if isinstance(metrics_data, dict) else {}
    rows = []

    values = payload.get("values")
    if isinstance(values, list):
        for item in values:
            if not isinstance(item, dict):
                continue
            extra = {k: v for k, v in item.items() if k not in {"name", "value", "startDate", "endDate", "currency"}}
            rows.append(
                {
                    "metric_name": item.get("name"),
                    "metric_value": item.get("value"),
                    "metric_start_date": item.get("startDate"),
                    "metric_end_date": item.get("endDate"),
                    "metric_currency": item.get("currency"),
                    "metric_extra": extra,
                }
            )
        return rows

    if "value" in payload:
        rows.append(
            {
                "metric_name": payload.get("name"),
                "metric_value": payload.get("value"),
                "metric_start_date": payload.get("startDate"),
                "metric_end_date": payload.get("endDate"),
                "metric_currency": payload.get("currency"),
                "metric_extra": {},
            }
        )
    return rows


def build_site_kpi_rows(metrics_data, project_id, device, segment_id=None):
    rows = []
    for metric in metrics_response_to_rows(metrics_data):
        rows.append(
            {
                "project_id": project_id,
                "device": device,
                "segment_id": segment_id,
                **metric,
            }
        )
    return rows


def build_group_kpi_rows(
    endpoint,
    token,
    project_id,
    start_date,
    end_date,
    page_groups,
    device,
    goal_ids=None,
    segment_ids=None,
):
    goal_ids = goal_ids or []
    segment_ids = segment_ids or [None]
    rows = []
    for page_group in page_groups:
        page_group_id = page_group.get("id")
        if page_group_id is None:
            continue

        for segment_id in segment_ids:
            segment_filter = [segment_id] if segment_id is not None else None

            base_metrics = get_page_group_metrics(
                endpoint,
                token,
                project_id,
                page_group_id,
                start_date,
                end_date,
                device=device,
                segment_ids=segment_filter,
            )
            for metric in metrics_response_to_rows(base_metrics):
                rows.append(
                    {
                        "project_id": project_id,
                        "device": device,
                        "segment_id": segment_id,
                        "mapping_id": page_group.get("mapping_id"),
                        "mapping_name": page_group.get("mapping_name"),
                        "page_group_id": page_group_id,
                        "page_group_name": page_group.get("name"),
                        "page_group_category": page_group.get("category"),
                        "goal_id": None,
                        **metric,
                    }
                )

            web_vitals = get_page_group_web_vitals(
                endpoint,
                token,
                project_id,
                page_group_id,
                start_date,
                end_date,
                device=device,
                segment_ids=segment_filter,
            )
            for metric in metrics_response_to_rows(web_vitals):
                rows.append(
                    {
                        "project_id": project_id,
                        "device": device,
                        "segment_id": segment_id,
                        "mapping_id": page_group.get("mapping_id"),
                        "mapping_name": page_group.get("mapping_name"),
                        "page_group_id": page_group_id,
                        "page_group_name": page_group.get("name"),
                        "page_group_category": page_group.get("category"),
                        "goal_id": None,
                        **metric,
                    }
                )

            # Goal-scoped metrics are intentionally added only at page-group level.
            for goal_id in goal_ids:
                goal_conv_rate = get_page_group_conversion_rate(
                    endpoint,
                    token,
                    project_id,
                    page_group_id,
                    start_date,
                    end_date,
                    goal_id=goal_id,
                    device=device,
                    segment_ids=segment_filter,
                )
                for metric in metrics_response_to_rows(goal_conv_rate):
                    rows.append(
                        {
                            "project_id": project_id,
                            "device": device,
                            "segment_id": segment_id,
                            "mapping_id": page_group.get("mapping_id"),
                            "mapping_name": page_group.get("mapping_name"),
                            "page_group_id": page_group_id,
                            "page_group_name": page_group.get("name"),
                            "page_group_category": page_group.get("category"),
                            "goal_id": goal_id,
                            **metric,
                        }
                    )
    return rows


def is_numeric(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def segment_value_column_name(segment_id):
    if segment_id is None:
        return "metric_value_all"
    return f"metric_value_segment_{segment_id}"


def segment_delta_column_name(segment_id):
    if segment_id is None:
        return "delta_vs_ref_all"
    return f"delta_vs_ref_segment_{segment_id}"


def normalize_segment_order(segment_ids, rows):
    ordered = []
    for segment_id in segment_ids or []:
        if segment_id not in ordered:
            ordered.append(segment_id)

    for row in rows:
        segment_id = row.get("segment_id")
        if segment_id not in ordered:
            ordered.append(segment_id)

    if not ordered:
        ordered = [None]
    return ordered


def pivot_rows_by_segment(rows, key_fields, segment_order):
    grouped = {}

    for row in rows:
        normalized = dict(row)
        normalized["metric_extra_json"] = normalize_excel_value(normalized.get("metric_extra"))
        key = tuple(normalized.get(field) for field in key_fields)

        if key not in grouped:
            grouped[key] = {
                "base": {field: normalized.get(field) for field in key_fields},
                "values": {},
            }
        grouped[key]["values"][normalized.get("segment_id")] = normalized.get("metric_value")

    return list(grouped.values())


def apply_reference_coloring_on_pivot(ws, reference_segment_id):
    if reference_segment_id is None or ws.max_row < 2:
        return

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    reference_column = headers.get(segment_value_column_name(reference_segment_id))
    if not reference_column:
        return
    metric_name_column = headers.get("metric_name")

    compare_columns = [
        idx
        for name, idx in headers.items()
        if name.startswith("metric_value_segment_") and idx != reference_column
    ]
    delta_columns = [
        idx
        for name, idx in headers.items()
        if name.startswith("delta_vs_ref_")
    ]
    if not compare_columns:
        return

    fill_green = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    fill_red = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    fill_yellow = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")

    for row_idx in range(2, ws.max_row + 1):
        ref_value = ws.cell(row=row_idx, column=reference_column).value
        if not is_numeric(ref_value):
            continue
        metric_name = ws.cell(row=row_idx, column=metric_name_column).value if metric_name_column else None
        reverse_coloring = metric_name == "bounceRate"

        for col_idx in compare_columns + delta_columns:
            value_cell = ws.cell(row=row_idx, column=col_idx)
            value = value_cell.value
            if not is_numeric(value):
                continue

            # For delta columns, compare against 0; for value columns, compare against ref
            if col_idx in delta_columns:
                cmp_ref = 0
            else:
                cmp_ref = float(ref_value)

            if float(value) > cmp_ref:
                value_cell.fill = fill_red if reverse_coloring else fill_green
            elif float(value) < cmp_ref:
                value_cell.fill = fill_green if reverse_coloring else fill_red
            else:
                value_cell.fill = fill_yellow


def build_segment_headers_with_deltas(segment_order, reference_segment_id):
    headers = []
    for seg_id in segment_order:
        headers.append(segment_value_column_name(seg_id))
        if reference_segment_id is not None and seg_id != reference_segment_id:
            headers.append(segment_delta_column_name(seg_id))
    return headers


def build_segment_values_with_deltas(values, segment_order, reference_segment_id):
    cells = []
    for seg_id in segment_order:
        raw = values.get(seg_id)
        cells.append(normalize_excel_value(raw))
        if reference_segment_id is not None and seg_id != reference_segment_id:
            cells.append(None)  # placeholder, formula inserted later
    return cells


def insert_delta_formulas(ws, key_fields_count, segment_order, reference_segment_id):
    if reference_segment_id is None or ws.max_row < 2:
        return

    # Build a map of segment_id -> (value_col_index, delta_col_index)
    ref_col = None
    delta_cols = {}  # segment_id -> (value_col_1based, delta_col_1based)
    col = key_fields_count + 1  # 1-based
    for seg_id in segment_order:
        value_col = col
        col += 1
        if seg_id == reference_segment_id:
            ref_col = value_col
        else:
            if reference_segment_id is not None:
                delta_col = col
                col += 1
                delta_cols[seg_id] = (value_col, delta_col)

    if ref_col is None:
        return

    ref_letter = get_column_letter(ref_col)
    for seg_id, (val_col, delta_col) in delta_cols.items():
        val_letter = get_column_letter(val_col)
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=delta_col).value = (
                f"=({val_letter}{row_idx}-{ref_letter}{row_idx})/{ref_letter}{row_idx}"
            )


METRIC_NUMBER_FORMATS = {
    "bounceRate":           '0.00"%"',
    "cartAverage":          '#,##0.00 "€"',
    "pageviewAverage":      "0.00",
    "revenueSum":           '#,##0.00 "€"',
    "sessionTimeAverage":   "0.00",
    "conversionCount":      "#,##0",
    "conversionRate":       '0.00"%"',
    "visits":               "#,##0",
}

DELTA_NUMBER_FORMAT = "0.00%"


def apply_number_formatting(ws):
    if ws.max_row < 2:
        return

    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    metric_name_col = headers.get("metric_name")
    if not metric_name_col:
        return

    value_columns = [
        idx for name, idx in headers.items()
        if name.startswith("metric_value_")
    ]
    delta_columns = [
        idx for name, idx in headers.items()
        if name.startswith("delta_vs_ref_")
    ]

    for row_idx in range(2, ws.max_row + 1):
        metric_name = ws.cell(row=row_idx, column=metric_name_col).value
        fmt = METRIC_NUMBER_FORMATS.get(metric_name)
        if fmt:
            for col_idx in value_columns:
                ws.cell(row=row_idx, column=col_idx).number_format = fmt
        for col_idx in delta_columns:
            ws.cell(row=row_idx, column=col_idx).number_format = DELTA_NUMBER_FORMAT


def export_kpis_excel(export_dir, filename, site_rows, group_rows, reference_segment_id=None, segment_ids=None):
    export_path = Path(export_dir)
    export_path.mkdir(parents=True, exist_ok=True)
    file_path = export_path / filename

    wb = Workbook()
    segment_order = normalize_segment_order(segment_ids, site_rows + group_rows)
    segment_columns = build_segment_headers_with_deltas(segment_order, reference_segment_id)

    ws_site = wb.active
    ws_site.title = "site_wide_kpis"
    site_key_fields = [
        "project_id",
        "device",
        "metric_name",
        "metric_currency",
        "metric_extra_json",
    ]
    ws_site.append(site_key_fields + segment_columns)

    site_pivot = pivot_rows_by_segment(site_rows, site_key_fields, segment_order)
    for item in site_pivot:
        base = item["base"]
        values = item["values"]
        ws_site.append(
            [base.get(field) for field in site_key_fields]
            + build_segment_values_with_deltas(values, segment_order, reference_segment_id)
        )
    ws_site.freeze_panes = "A2"

    ws_group = wb.create_sheet("group_id_kpis")
    group_key_fields = [
        "project_id",
        "device",
        "mapping_id",
        "mapping_name",
        "page_group_id",
        "page_group_name",
        "page_group_category",
        "goal_id",
        "metric_name",
        "metric_currency",
        "metric_extra_json",
    ]
    ws_group.append(group_key_fields + segment_columns)

    group_pivot = pivot_rows_by_segment(group_rows, group_key_fields, segment_order)
    for item in group_pivot:
        base = item["base"]
        values = item["values"]
        ws_group.append(
            [base.get(field) for field in group_key_fields]
            + build_segment_values_with_deltas(values, segment_order, reference_segment_id)
        )
    ws_group.freeze_panes = "A2"

    insert_delta_formulas(ws_site, len(site_key_fields), segment_order, reference_segment_id)
    insert_delta_formulas(ws_group, len(group_key_fields), segment_order, reference_segment_id)

    apply_number_formatting(ws_site)
    apply_number_formatting(ws_group)

    apply_reference_coloring_on_pivot(ws_site, reference_segment_id)
    apply_reference_coloring_on_pivot(ws_group, reference_segment_id)

    wb.save(file_path)
    return file_path


def display_metrics(label, metrics_data, goal_conversions=None, goal_conversion_rate=None):
    sessions = extract_metric_value(metrics_data, "visits", fallback_key="sessions")
    pageviews = extract_metric_value(metrics_data, "pageviews", fallback_key="pageviews")
    pageview_average = extract_metric_value(metrics_data, "pageviewAverage")
    bounce_rate = extract_metric_value(metrics_data, "bounceRate", fallback_key="bounceRate")

    conversions = extract_metric_value(metrics_data, "conversionCount")
    if conversions is None and goal_conversions:
        conversions = extract_single_value(goal_conversions, preferred_name="conversionCount")

    conv_rate = extract_metric_value(metrics_data, "conversionRate")
    if conv_rate is None and goal_conversion_rate:
        conv_rate = extract_single_value(goal_conversion_rate, preferred_name="conversionRate")

    print(f"\n{label}:")
    print(f"  Sessions/visites:   {format_count(sessions)}")
    if isinstance(pageviews, (int, float)):
        print(f"  Pageviews:          {format_count(pageviews)}")
    elif isinstance(pageview_average, (int, float)):
        print(f"  Pages / session:    {pageview_average:.2f}")
    else:
        print("  Pageviews:          N/A")
    print(f"  Taux de rebond:     {format_percentage(bounce_rate)}")
    if isinstance(conversions, (int, float)):
        print(f"  Conversions:        {format_count(conversions)}")
    if isinstance(conv_rate, (int, float)):
        print(f"  Taux de conversion: {format_percentage(conv_rate)}")


def main():
    print("="*70)
    print("📊 CONTENTSQUARE - CONVERSIONS E-COMMERCE")
    print("="*70)
    print()
    print("🔧 Configuration:")
    print(f"   Segments: {SEGMENT_IDS_TO_ANALYZE}")
    print(f"   By device: {ANALYZE_BY_DEVICE}")
    print(f"   Période: {START_DATE} -> {END_DATE}")
    print(f"   Goal ID: {GOAL_ID}")
    print(f"   Goal IDs (group KPI only): {GOAL_IDS}")
    print(f"   Page-group ID (KPI scope): {PAGE_GROUP_ID}")
    print(f"   Page-group mapping ID (export): {PAGE_GROUP_MAPPING_ID}")
    print(f"   Export dir: {EXPORT_DIR}")
    print(f"   KPI Excel: {KPI_EXCEL_FILENAME}")
    print()

    # 1. Authentification
    try:
        token, endpoint = get_token(CLIENT_ID, CLIENT_SECRET, PROJECT_ID)
        print("✅ Token généré avec succès !")
        print(f"📡 Endpoint: {endpoint}\n")
    except Exception as e:
        print(f"❌ Erreur d'authentification: {e}")
        return

    # 2. Dates
    try:
        start_date = parse_config_date(START_DATE, "START_DATE").replace(
            hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc
        )
        end_date = parse_config_date(END_DATE, "END_DATE").replace(
            hour=23, minute=59, second=59, microsecond=999000, tzinfo=timezone.utc
        )
    except ValueError as e:
        print(f"❌ Erreur configuration dates: {e}")
        return

    if start_date > end_date:
        print("❌ Erreur configuration dates: START_DATE doit être antérieure ou égale à END_DATE.")
        return

    start_date_iso = start_date.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    end_date_iso = end_date.strftime("%Y-%m-%dT%H:%M:%S.999Z")
    print(f"📅 Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}\n")

    # 3. Segments
    try:
        segments_data = get_segments(endpoint, token, PROJECT_ID)
        segments = segments_data.get("payload", [])
        print(f"✅ {len(segments)} segment(s) disponible(s)")
        if segments:
            print("\n📋 Segments disponibles:")
            for seg in segments:
                print(f"  - ID: {seg.get('id')} → {seg.get('name')}")
        print()
    except Exception as e:
        print(f"⚠️ Impossible de lister les segments: {e}")
        segments = []

    segment_rows = [{"id": s.get("id"), "name": s.get("name"), "extra": ""} for s in segments]
    segment_file = export_ids_file(EXPORT_DIR, "segments_ids", "Segment IDs", segment_rows)
    print(f"📝 Segment IDs exportés: {segment_file}")

    # 3bis. Goals
    try:
        goals_data = get_goals(endpoint, token, PROJECT_ID)
        goals = goals_data.get("payload", [])
        print(f"✅ {len(goals)} goal(s) disponible(s)")
    except Exception as e:
        print(f"⚠️ Impossible de lister les goals: {e}")
        goals = []

    goal_rows = [{"id": g.get("id"), "name": g.get("name"), "extra": g.get("type", "")} for g in goals]
    goals_file = export_ids_file(EXPORT_DIR, "goals_ids", "Goal IDs", goal_rows)
    print(f"📝 Goal IDs exportés: {goals_file}")

    # 3ter. Page groups
    try:
        if PAGE_GROUP_ID is not None:
            selected_page_group = find_page_group_by_id(endpoint, token, PROJECT_ID, PAGE_GROUP_ID)
            if not selected_page_group:
                print(f"⚠️ Page group ID {PAGE_GROUP_ID} introuvable.")
                page_groups = []
            else:
                page_groups = [selected_page_group]
                print(f"✅ Export KPI limité au page group ID {PAGE_GROUP_ID} ({selected_page_group.get('name')})")
        else:
            page_groups = get_page_groups_for_mapping(endpoint, token, PROJECT_ID, PAGE_GROUP_MAPPING_ID)
            if page_groups:
                print(f"✅ {len(page_groups)} page group(s) disponible(s) pour mapping {PAGE_GROUP_MAPPING_ID}")
            else:
                guessed_page_group = find_page_group_by_id(endpoint, token, PROJECT_ID, PAGE_GROUP_MAPPING_ID)
                if guessed_page_group:
                    page_groups = [guessed_page_group]
                    print(
                        f"⚠️ {PAGE_GROUP_MAPPING_ID} ressemble à un page-group ID (pas un mapping ID). "
                        "Export KPI limité à ce page group."
                    )
                else:
                    print(
                        f"⚠️ Aucun page group trouvé pour mapping {PAGE_GROUP_MAPPING_ID}. "
                        "Fallback: récupération de tous les mappings."
                    )
                    page_groups = get_all_page_groups(endpoint, token, PROJECT_ID)
                    print(f"✅ {len(page_groups)} page group(s) disponible(s) sur tous les mappings")
    except Exception as e:
        print(f"⚠️ Impossible de lister les page groups pour mapping {PAGE_GROUP_MAPPING_ID}: {e}")
        page_groups = []

    page_group_rows = [
        {
            "id": pg.get("id"),
            "name": pg.get("name"),
            "extra": (
                f"mappingId={pg.get('mapping_id', '')};"
                f"mapping={pg.get('mapping_name', '')};"
                f"category={pg.get('category', '')}"
            ),
        }
        for pg in page_groups
    ]
    page_groups_file = export_ids_file(EXPORT_DIR, "page_group_ids", "Page Group IDs", page_group_rows)
    print(f"📝 Page Group IDs exportés: {page_groups_file}")
    print()

    # 4. Métriques globales
    print("-"*70)
    print("📊 MÉTRIQUES GLOBALES (Tous visiteurs)")
    print("-"*70)
    site_kpi_rows = []
    group_kpi_rows = []
    segment_scope = SEGMENT_IDS_TO_ANALYZE[:] if SEGMENT_IDS_TO_ANALYZE else [None]
    device_scope = get_analysis_devices(ANALYZE_BY_DEVICE)
    detailed_device_scope = [device for device in device_scope if device != "all"]
    reference_segment_id = SEGMENT_IDS_TO_ANALYZE[0] if SEGMENT_IDS_TO_ANALYZE else None
    try:
        for device in device_scope:
            site_metrics = get_site_metrics(
                endpoint,
                token,
                PROJECT_ID,
                start_date_iso,
                end_date_iso,
                device=device,
            )
            for segment_id in segment_scope:
                segment_filter = [segment_id] if segment_id is not None else None
                site_metrics_for_segment = get_site_metrics(
                    endpoint,
                    token,
                    PROJECT_ID,
                    start_date_iso,
                    end_date_iso,
                    device=device,
                    segment_ids=segment_filter,
                )
                site_kpi_rows.extend(
                    build_site_kpi_rows(
                        site_metrics_for_segment,
                        PROJECT_ID,
                        device=device,
                        segment_id=segment_id,
                    )
                )
            if device == "all":
                goal_conversions = get_ecommerce_conversions(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso)
                goal_conv_rate = get_ecommerce_conversion_rate(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso)
                display_metrics("ALL DEVICES", site_metrics, goal_conversions, goal_conv_rate)
    except Exception as e:
        print(f"❌ Erreur: {e}")
    print()

    # 5. Métriques par device
    if detailed_device_scope:
        print("-"*70)
        print("📱 MÉTRIQUES PAR DEVICE")
        print("-"*70)
        for device in detailed_device_scope:
            try:
                site_metrics = get_site_metrics(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso, device=device)
                goal_conversions = get_ecommerce_conversions(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso, device=device)
                goal_conv_rate = get_ecommerce_conversion_rate(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso, device=device)
                display_metrics(device.upper(), site_metrics, goal_conversions, goal_conv_rate)
            except Exception as e:
                print(f"\n{device.upper()}: ❌ Erreur - {e}")
        print()

    print("-"*70)
    print("📁 EXPORT KPI EXCEL")
    print("-"*70)
    try:
        for device in device_scope:
            group_kpi_rows.extend(
                build_group_kpi_rows(
                    endpoint,
                    token,
                    PROJECT_ID,
                    start_date_iso,
                    end_date_iso,
                    page_groups,
                    device=device,
                    goal_ids=GOAL_IDS,
                    segment_ids=segment_scope,
                )
            )
        excel_path = export_kpis_excel(
            EXPORT_DIR,
            KPI_EXCEL_FILENAME,
            site_kpi_rows,
            group_kpi_rows,
            reference_segment_id=reference_segment_id,
            segment_ids=segment_scope,
        )
        print(f"📝 KPI Excel exporté: {excel_path}")
        print(f"   - Site-wide KPIs: {len(site_kpi_rows)} ligne(s)")
        print(f"   - Group-ID KPIs:  {len(group_kpi_rows)} ligne(s)")
    except Exception as e:
        print(f"❌ Erreur export KPI Excel: {e}")
    print()

    # 6. Métriques par segment
    if SEGMENT_IDS_TO_ANALYZE:
        print("-"*70)
        print("📊 MÉTRIQUES PAR SEGMENT")
        print("-"*70)
        for seg_id in SEGMENT_IDS_TO_ANALYZE:
            seg_name = f"Segment {seg_id}"
            for seg in segments:
                if seg.get("id") == seg_id:
                    seg_name = seg.get("name")
                    break
            try:
                site_metrics = get_site_metrics(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso, segment_ids=[seg_id])
                goal_conversions = get_ecommerce_conversions(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso, segment_ids=[seg_id])
                goal_conv_rate = get_ecommerce_conversion_rate(endpoint, token, PROJECT_ID, start_date_iso, end_date_iso, segment_ids=[seg_id])
                display_metrics(seg_name, site_metrics, goal_conversions, goal_conv_rate)
            except Exception as e:
                print(f"\n{seg_name}: ❌ Erreur - {e}")
        print()

    print("="*70)
    print("✅ ANALYSE TERMINÉE")
    print("="*70)
    print()
    print("💡 Pour analyser des segments spécifiques, copie les IDs ci-dessus et ajoute-les dans SEGMENT_IDS_TO_ANALYZE")
    print(f"📅 Période: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}\n")


if __name__ == "__main__":
    main()
